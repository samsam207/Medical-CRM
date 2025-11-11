from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required
from app import db, cache
from app.models.payment import Payment, PaymentMethod, PaymentStatus
from app.models.visit import Visit, VisitStatus
from app.models.patient import Patient
from app.utils.decorators import receptionist_required, validate_json, log_audit
from app.utils.validators import validate_payment_amount
from app.services.payment_service import PaymentService
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO

payments_bp = Blueprint('payments', __name__)

@payments_bp.route('', methods=['GET'])
@jwt_required()
def get_payments():
    """Get payments with optional filters"""
    try:
        
        patient_id = request.args.get('patient_id', type=int)
        visit_id = request.args.get('visit_id', type=int)
        clinic_id = request.args.get('clinic_id', type=int)
        doctor_id = request.args.get('doctor_id', type=int)
        status = request.args.get('status')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        date = request.args.get('date')  # Single date filter
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 50, type=int)
        
        # Join with Visit table for clinic and doctor filtering
        query = Payment.query.join(Visit, Payment.visit_id == Visit.id)
        
        if patient_id:
            query = query.filter(Payment.patient_id == patient_id)
        if visit_id:
            query = query.filter(Payment.visit_id == visit_id)
        if clinic_id:
            query = query.filter(Visit.clinic_id == clinic_id)
        if doctor_id:
            query = query.filter(Visit.doctor_id == doctor_id)
        if status:
            try:
                status_enum = PaymentStatus(status)
                query = query.filter(Payment.status == status_enum)
            except ValueError:
                return jsonify({'message': 'Invalid status'}), 400
        if date:
            try:
                date_obj = datetime.strptime(date, '%Y-%m-%d').date()
                query = query.filter(db.func.date(Payment.created_at) == date_obj)
            except ValueError:
                return jsonify({'message': 'Invalid date format. Use YYYY-MM-DD'}), 400
        elif start_date:
            try:
                start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
                query = query.filter(db.func.date(Payment.created_at) >= start_date_obj)
            except ValueError:
                return jsonify({'message': 'Invalid start_date format. Use YYYY-MM-DD'}), 400
        if end_date:
            try:
                end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
                query = query.filter(db.func.date(Payment.created_at) <= end_date_obj)
            except ValueError:
                return jsonify({'message': 'Invalid end_date format. Use YYYY-MM-DD'}), 400
        
        # Order by creation date
        query = query.order_by(Payment.created_at.desc())
        
        # Paginate
        payments = query.paginate(
            page=page, per_page=per_page, error_out=False
        )
        
        # Serialize with visit data
        def serialize_payment(payment):
            try:
                payment_dict = payment.to_dict()
                
                # Add visit data if available
                try:
                    if hasattr(payment, 'visit') and payment.visit:
                        payment_dict['visit'] = {
                            'id': payment.visit.id
                        }
                        
                        # Manually query for related data to avoid relationship recursion
                        from app.models.patient import Patient
                        from app.models.doctor import Doctor
                        from app.models.clinic import Clinic
                        from app.models.service import Service
                        
                        patient = db.session.get(Patient, payment.visit.patient_id) if payment.visit.patient_id else None
                        doctor = db.session.get(Doctor, payment.visit.doctor_id) if payment.visit.doctor_id else None
                        clinic = db.session.get(Clinic, payment.visit.clinic_id) if payment.visit.clinic_id else None
                        service = db.session.get(Service, payment.visit.service_id) if payment.visit.service_id else None
                        
                        payment_dict['visit']['patient'] = {
                            'id': patient.id,
                            'name': patient.name,
                            'phone': patient.phone
                        } if patient else None
                        
                        payment_dict['visit']['doctor'] = {
                            'id': doctor.id,
                            'name': doctor.name
                        } if doctor else None
                        
                        payment_dict['visit']['clinic'] = {
                            'id': clinic.id,
                            'name': clinic.name
                        } if clinic else None
                        
                        payment_dict['visit']['service'] = {
                            'id': service.id,
                            'name': service.name,
                            'price': float(service.price) if service.price else None
                        } if service else None
                except Exception as e:
                    payment_dict['visit'] = None
                    
                return payment_dict
            except Exception as e:
                # Fallback to basic serialization
                return {
                    'id': payment.id,
                    'visit_id': payment.visit_id,
                    'patient_id': payment.patient_id,
                    'total_amount': float(payment.total_amount),
                    'amount_paid': float(payment.amount_paid),
                    'discount_amount': float(payment.discount_amount),
                    'remaining_amount': payment.remaining_amount,
                    'payment_method': payment.payment_method.value if payment.payment_method else None,
                    'status': payment.status.value if payment.status else None,
                    'doctor_share': float(payment.doctor_share),
                    'center_share': float(payment.center_share),
                    'paid_at': payment.paid_at.isoformat() if payment.paid_at else None,
                    'created_at': payment.created_at.isoformat() if payment.created_at else None,
                    'visit': None
                }
        
        return jsonify({
            'payments': [serialize_payment(payment) for payment in payments.items],
            'total': payments.total,
            'pages': payments.pages,
            'current_page': page,
            'per_page': per_page
        }), 200
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'message': f'An error occurred: {str(e)}'}), 500

@payments_bp.route('/<int:payment_id>', methods=['GET'])
@jwt_required()
def get_payment(payment_id):
    """Get payment details"""
    payment = Payment.query.get_or_404(payment_id)
    return jsonify({'payment': payment.to_dict()}), 200

@payments_bp.route('/<int:payment_id>/process', methods=['PUT'])
@receptionist_required
@validate_json(['amount_paid', 'payment_method'])
@log_audit('process_existing_payment', 'payment')
def process_existing_payment(payment_id, data, current_user):
    """Process an existing pending payment"""
    payment = Payment.query.get_or_404(payment_id)
    
    # Check if payment is pending or partially paid
    if payment.status not in [PaymentStatus.PENDING, PaymentStatus.PARTIALLY_PAID, PaymentStatus.APPOINTMENT_COMPLETED]:
        return jsonify({'message': 'Only pending, partially paid, or appointment completed payments can be processed'}), 400
    
    # Validate payment method
    try:
        payment_method_enum = PaymentMethod(data['payment_method'])
    except ValueError:
        return jsonify({'message': 'Invalid payment method'}), 400
    
    # Get discount amount if provided
    discount_amount = float(data.get('discount_amount', 0))
    if discount_amount < 0:
        return jsonify({'message': 'Discount amount cannot be negative'}), 400
    
    if discount_amount > payment.total_amount:
        return jsonify({'message': 'Discount amount cannot exceed total amount'}), 400
    
    # Get amount to be set as total paid (frontend sends the full amount to set)
    new_total_paid = float(data['amount_paid'])
    if new_total_paid <= 0:
        return jsonify({'message': 'Payment amount must be greater than 0'}), 400
    
    # Validate total paid doesn't exceed total amount minus discount
    max_allowed = float(payment.total_amount) - discount_amount
    if new_total_paid > max_allowed:
        return jsonify({'message': f'Total payment cannot exceed ${max_allowed:.2f} after discount'}), 400
    
    # Update payment with discount if this is the first payment
    if payment.status == PaymentStatus.PENDING:
        payment.discount_amount = discount_amount
    
    # Get visit before updates
    visit = payment.visit
    
    # Update payment amounts
    payment.amount_paid = new_total_paid
    payment.payment_method = payment_method_enum
    payment.paid_at = datetime.utcnow()
    
    # Determine payment status
    remaining = float(payment.total_amount) - discount_amount - new_total_paid
    if remaining <= 0.01:  # Account for floating point precision
        payment.status = PaymentStatus.PAID
        # Update visit status to completed
        if visit:
            visit.status = VisitStatus.COMPLETED
            visit.end_time = datetime.utcnow()
    else:
        payment.status = PaymentStatus.PARTIALLY_PAID
    
    db.session.commit()
    
    # Emit real-time update for payment processing
    from app import socketio
    from app.services.queue_service import QueueService
    
    # Emit payment processed event
    socketio.emit('payment_processed', {
        'payment': payment.to_dict(),
        'visit': visit.to_dict() if visit else None
    })
    
    # Emit queue updates if visit exists
    if visit:
        queue_service = QueueService()
        queue_data = queue_service.get_clinic_queue(visit.clinic_id)
        socketio.emit('queue_updated', queue_data, room=f'clinic_{visit.clinic_id}')
        
        doctor_queue_data = queue_service.get_doctor_queue(visit.doctor_id)
        socketio.emit('queue_updated', doctor_queue_data, room=f'doctor_{visit.doctor_id}')
    
    # Invalidate cache
    cache.clear()
    
    return jsonify({
        'message': 'Payment processed successfully',
        'payment': payment.to_dict()
    }), 200

@payments_bp.route('', methods=['POST'])
@receptionist_required
@validate_json(['visit_id', 'payment_method', 'amount_paid'])
@log_audit('process_payment', 'payment')
def process_payment(data, current_user):
    """Process payment for visit"""
    visit_id = data['visit_id']
    payment_method = data['payment_method']
    amount_paid = data['amount_paid']
    
    # Get visit
    visit = Visit.query.get_or_404(visit_id)
    
    # Check if visit is ready for payment
    if visit.status != VisitStatus.PENDING_PAYMENT:
        return jsonify({'message': 'Visit is not ready for payment'}), 400
    
    # Validate payment amount
    is_valid, message = validate_payment_amount(visit_id, amount_paid)
    if not is_valid:
        return jsonify({'message': message}), 400
    
    # Validate payment method
    try:
        payment_method_enum = PaymentMethod(payment_method)
    except ValueError:
        return jsonify({'message': 'Invalid payment method'}), 400
    
    # Process payment using service
    payment_service = PaymentService()
    payment = payment_service.process_payment(
        visit_id=visit_id,
        payment_method=payment_method_enum,
        amount_paid=amount_paid
    )
    
    # Update visit status
    visit.status = VisitStatus.COMPLETED
    visit.end_time = datetime.utcnow()
    
    db.session.commit()
    
    return jsonify({
        'message': 'Payment processed successfully',
        'payment': payment.to_dict(),
        'invoice': payment_service.generate_invoice_data(payment.id)
    }), 201

@payments_bp.route('/invoice/<int:visit_id>', methods=['GET'])
@jwt_required()
def get_invoice(visit_id):
    """Generate printable invoice for visit"""
    visit = Visit.query.get_or_404(visit_id)
    
    if not visit.payment:
        return jsonify({'message': 'No payment found for this visit'}), 404
    
    payment_service = PaymentService()
    invoice_data = payment_service.generate_invoice_data(visit.payment.id)
    
    return jsonify({'invoice': invoice_data}), 200

@payments_bp.route('/refund/<int:payment_id>', methods=['POST'])
@receptionist_required
@log_audit('refund_payment', 'payment')
def refund_payment(payment_id, current_user):
    """Refund payment"""
    payment = Payment.query.get_or_404(payment_id)
    
    if payment.status != PaymentStatus.PAID:
        return jsonify({'message': 'Only paid payments can be refunded'}), 400
    
    # Update payment status
    payment.status = PaymentStatus.REFUNDED
    
    # Update visit status back to pending payment
    visit = Visit.query.get(payment.visit_id)
    if visit:
        visit.status = VisitStatus.PENDING_PAYMENT
    
    db.session.commit()
    
    return jsonify({
        'message': 'Payment refunded successfully',
        'payment': payment.to_dict()
    }), 200

@payments_bp.route('/statistics', methods=['GET'])
@jwt_required()
def get_payment_statistics():
    """Get payment statistics"""
    try:
        from app.models.visit import Visit
        clinic_id = request.args.get('clinic_id', type=int)
        doctor_id = request.args.get('doctor_id', type=int)
        
        # Validate clinic_id and doctor_id
        # request.args.get('clinic_id', type=int) returns None if not provided or invalid
        # So we only need to check if it's a valid positive integer
        if clinic_id is not None and clinic_id <= 0:
            clinic_id = None
        if doctor_id is not None and doctor_id <= 0:
            doctor_id = None
        
        # Build base query with filters
        base_query = Payment.query
        if clinic_id or doctor_id:
            # Join with visits to filter by clinic/doctor (payments must have visits for filtering)
            base_query = base_query.join(Visit, Payment.visit_id == Visit.id)
            if clinic_id:
                base_query = base_query.filter(Visit.clinic_id == clinic_id)
            if doctor_id:
                base_query = base_query.filter(Visit.doctor_id == doctor_id)
        
        # Get total payments
        total_payments = base_query.count()
        
        # Count by status
        pending_count = base_query.filter_by(status=PaymentStatus.PENDING).count()
        partially_paid_count = base_query.filter_by(status=PaymentStatus.PARTIALLY_PAID).count()
        paid_count = base_query.filter_by(status=PaymentStatus.PAID).count()
        refunded_count = base_query.filter_by(status=PaymentStatus.REFUNDED).count()
        
        # Count by method
        cash_count = base_query.filter_by(payment_method=PaymentMethod.CASH).count()
        visa_count = base_query.filter_by(payment_method=PaymentMethod.VISA).count()
        bank_transfer_count = base_query.filter_by(payment_method=PaymentMethod.BANK_TRANSFER).count()
        
        # Calculate totals - need to rebuild query for aggregations
        revenue_query = db.session.query(db.func.sum(Payment.amount_paid))
        if clinic_id or doctor_id:
            revenue_query = revenue_query.join(Visit, Payment.visit_id == Visit.id)
            if clinic_id:
                revenue_query = revenue_query.filter(Visit.clinic_id == clinic_id)
            if doctor_id:
                revenue_query = revenue_query.filter(Visit.doctor_id == doctor_id)
        total_revenue = revenue_query.filter(Payment.status == PaymentStatus.PAID).scalar() or 0
        
        # Total refunds
        refunds_query = db.session.query(db.func.sum(Payment.amount_paid))
        if clinic_id or doctor_id:
            refunds_query = refunds_query.join(Visit, Payment.visit_id == Visit.id)
            if clinic_id:
                refunds_query = refunds_query.filter(Visit.clinic_id == clinic_id)
            if doctor_id:
                refunds_query = refunds_query.filter(Visit.doctor_id == doctor_id)
        total_refunds = refunds_query.filter(Payment.status == PaymentStatus.REFUNDED).scalar() or 0
        
        # Pending amount
        pending_query = db.session.query(db.func.sum(Payment.total_amount - Payment.amount_paid))
        if clinic_id or doctor_id:
            pending_query = pending_query.join(Visit, Payment.visit_id == Visit.id)
            if clinic_id:
                pending_query = pending_query.filter(Visit.clinic_id == clinic_id)
            if doctor_id:
                pending_query = pending_query.filter(Visit.doctor_id == doctor_id)
        pending_amount = pending_query.filter(
            Payment.status.in_([PaymentStatus.PENDING, PaymentStatus.PARTIALLY_PAID])
        ).scalar() or 0
        
        return jsonify({
            'total': total_payments,
            'total_payments': total_payments,  # Alias for frontend compatibility
            'total_revenue': float(total_revenue),  # Alias for frontend compatibility
            'pending_count': pending_count,  # Alias for frontend compatibility
            'paid_count': paid_count,  # Alias for frontend compatibility
            'pending_amount': float(pending_amount),  # Alias for frontend compatibility
            'by_status': {
                'pending': pending_count,
                'partially_paid': partially_paid_count,
                'paid': paid_count,
                'refunded': refunded_count
            },
            'by_method': {
                'cash': cash_count,
                'visa': visa_count,
                'bank_transfer': bank_transfer_count
            },
            'totals': {
                'revenue': float(total_revenue),
                'refunds': float(total_refunds),
                'pending_amount': float(pending_amount)
            }
        }), 200
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'message': f'Error retrieving statistics: {str(e)}'}), 500

@payments_bp.route('/export', methods=['GET'])
@jwt_required()
def export_payments():
    """Export payments to Excel file"""
    
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    date = request.args.get('date')  # Single date for convenience
    clinic_id = request.args.get('clinic_id', type=int)
    doctor_id = request.args.get('doctor_id', type=int)
    status = request.args.get('status')
    method = request.args.get('method')
    
    # Join with Visit table for clinic and doctor filtering
    query = Payment.query.join(Visit, Payment.visit_id == Visit.id)
    
    # Filter by clinic
    if clinic_id:
        query = query.filter(Visit.clinic_id == clinic_id)
    
    # Filter by doctor
    if doctor_id:
        query = query.filter(Visit.doctor_id == doctor_id)
    
    # Filter by status
    if status:
        try:
            status_enum = PaymentStatus(status)
            query = query.filter(Payment.status == status_enum)
        except ValueError:
            return jsonify({'message': 'Invalid status'}), 400
    
    # Filter by method
    if method:
        try:
            method_enum = PaymentMethod(method)
            query = query.filter(Payment.payment_method == method_enum)
        except ValueError:
            return jsonify({'message': 'Invalid payment method'}), 400
    
    # Handle date range or single date
    if date:
        try:
            date_obj = datetime.strptime(date, '%Y-%m-%d').date()
            query = query.filter(db.func.date(Payment.created_at) == date_obj)
        except ValueError:
            return jsonify({'message': 'Invalid date format. Use YYYY-MM-DD'}), 400
    elif start_date or end_date:
        if start_date:
            try:
                start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
                query = query.filter(db.func.date(Payment.created_at) >= start_date_obj)
            except ValueError:
                return jsonify({'message': 'Invalid start_date format. Use YYYY-MM-DD'}), 400
        if end_date:
            try:
                end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
                query = query.filter(db.func.date(Payment.created_at) <= end_date_obj)
            except ValueError:
                return jsonify({'message': 'Invalid end_date format. Use YYYY-MM-DD'}), 400
    
    # Get payments with related data (already joined Visit above, now eager load relationships)
    payments = query.options(
        db.joinedload(Payment.visit).joinedload(Visit.patient),
        db.joinedload(Payment.visit).joinedload(Visit.doctor),
        db.joinedload(Payment.visit).joinedload(Visit.clinic),
        db.joinedload(Payment.visit).joinedload(Visit.service)
    ).all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Payments"
    
    # Define header styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Write headers
    headers = [
        'Payment ID', 'Date', 'Patient Name', 'Phone', 'Doctor', 'Clinic', 'Service',
        'Total Amount', 'Discount', 'Amount Paid', 'Remaining', 'Payment Method', 'Status'
    ]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Write data
    for row_num, payment in enumerate(payments, 2):
        visit = payment.visit
        patient = visit.patient if visit else None
        doctor = visit.doctor if visit else None
        clinic = visit.clinic if visit else None
        service = visit.service if visit else None
        
        ws.cell(row=row_num, column=1, value=payment.id)
        ws.cell(row=row_num, column=2, value=payment.created_at.strftime('%Y-%m-%d %H:%M') if payment.created_at else '')
        ws.cell(row=row_num, column=3, value=patient.name if patient else '')
        ws.cell(row=row_num, column=4, value=patient.phone if patient else '')
        ws.cell(row=row_num, column=5, value=f"Dr. {doctor.name}" if doctor else '')
        ws.cell(row=row_num, column=6, value=clinic.name if clinic else '')
        ws.cell(row=row_num, column=7, value=service.name if service else '')
        ws.cell(row=row_num, column=8, value=float(payment.total_amount))
        ws.cell(row=row_num, column=9, value=float(payment.discount_amount))
        ws.cell(row=row_num, column=10, value=float(payment.amount_paid))
        ws.cell(row=row_num, column=11, value=payment.remaining_amount)
        ws.cell(row=row_num, column=12, value=payment.payment_method.value if payment.payment_method else '')
        ws.cell(row=row_num, column=13, value=payment.status.value if payment.status else '')
    
    # Add total row
    total_row = row_num + 1
    ws.cell(row=total_row, column=7, value='TOTAL')
    ws.cell(row=total_row, column=8, value=f'=SUM(H2:H{row_num})')  # Total Amount
    ws.cell(row=total_row, column=9, value=f'=SUM(I2:I{row_num})')  # Discount
    ws.cell(row=total_row, column=10, value=f'=SUM(J2:J{row_num})')  # Amount Paid
    ws.cell(row=total_row, column=11, value=f'=SUM(K2:K{row_num})')  # Remaining
    
    # Style total row
    for col in [7, 8, 9, 10, 11]:
        cell = ws.cell(row=total_row, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # Adjust column widths
    column_widths = [12, 18, 20, 15, 20, 20, 20, 14, 12, 12, 12, 15, 15]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col_num)].width = width
    
    # Create in-memory file
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Generate filename
    if date:
        filename = f"payments_{date}.xlsx"
    elif start_date and end_date:
        filename = f"payments_{start_date}_to_{end_date}.xlsx"
    else:
        filename = f"payments_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
